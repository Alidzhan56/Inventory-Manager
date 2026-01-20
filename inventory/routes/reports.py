# inventory/routes/reports.py
from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy import func

from inventory.extensions import db
from inventory.models import Transaction, TransactionItem, Warehouse, Partner, Product, Stock
from inventory.utils.decorators import roles_required
from inventory.utils.translations import _

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


bp = Blueprint("reports", __name__, url_prefix="/reports")


def _get_owner_id():
    if current_user.role == "Developer":
        return None
    if current_user.role == "Admin / Owner":
        return current_user.id
    return current_user.created_by_id


def _require_owner_context(owner_id):
    if owner_id is None:
        flash(_("Owner context required."), "warning")
        return False
    return True


@bp.route("/", methods=["GET"])
@login_required
@roles_required("Admin / Owner", "Warehouse Manager", "Developer")
def reports_home():
    owner_id = _get_owner_id()

    # dropdowns for filters
    partners = Partner.query.order_by(Partner.name.asc())
    warehouses = Warehouse.query.order_by(Warehouse.name.asc())

    if owner_id is not None:
        partners = partners.filter_by(owner_id=owner_id)
        warehouses = warehouses.filter_by(owner_id=owner_id)

    return render_template(
        "reports.html",
        partners=partners.all(),
        warehouses=warehouses.all(),
    )


def _parse_date(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except Exception:
        return None


def _transactions_query(owner_id):
    q = (
        Transaction.query
        .join(Warehouse, Transaction.warehouse_id == Warehouse.id)
        .options(
            db.joinedload(Transaction.items).joinedload(TransactionItem.product),
            db.joinedload(Transaction.partner),
            db.joinedload(Transaction.warehouse),
            db.joinedload(Transaction.user),
        )
        .order_by(Transaction.date.desc())
    )
    if owner_id is not None:
        q = q.filter(Warehouse.owner_id == owner_id)
    return q


def _apply_tx_filters(q):
    ttype = (request.args.get("type") or "").strip()
    partner_id = request.args.get("partner_id", type=int)
    warehouse_id = request.args.get("warehouse_id", type=int)
    date_from = _parse_date(request.args.get("from") or "")
    date_to = _parse_date(request.args.get("to") or "")

    if ttype in {"Purchase", "Sale"}:
        q = q.filter(Transaction.type == ttype)

    if partner_id:
        q = q.filter(Transaction.partner_id == partner_id)

    if warehouse_id:
        q = q.filter(Transaction.warehouse_id == warehouse_id)

    if date_from:
        q = q.filter(Transaction.date >= date_from)

    if date_to:
        # include whole day
        q = q.filter(Transaction.date < (date_to.replace(hour=23, minute=59, second=59)))

    # Sales Agent safety rule: allow only Sales exports (if you want)
    if current_user.role == "Sales Agent":
        q = q.filter(Transaction.type == "Sale")

    return q


@bp.route("/transactions.xlsx")
@login_required
@roles_required("Admin / Owner", "Warehouse Manager", "Developer", "Sales Agent")
def transactions_xlsx():
    owner_id = _get_owner_id()
    if not _require_owner_context(owner_id):
        return redirect(url_for("reports.reports_home"))

    q = _apply_tx_filters(_transactions_query(owner_id))
    txns = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Type", "Partner", "Warehouse", "User", "Items", "Total"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for t in txns:
        total = sum((i.total_price or 0) for i in (t.items or []))
        ws.append([
            t.date.strftime("%Y-%m-%d %H:%M") if t.date else "",
            t.type or "",
            (t.partner.name if t.partner else ""),
            (t.warehouse.name if t.warehouse else ""),
            (t.user.username if t.user else ""),
            len(t.items or []),
            float(total),
        ])

    # Autosize-ish
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="transactions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/stock.xlsx")
@login_required
@roles_required("Admin / Owner", "Warehouse Manager", "Developer")
def stock_xlsx():
    owner_id = _get_owner_id()
    if not _require_owner_context(owner_id):
        return redirect(url_for("reports.reports_home"))

    q = (
        Stock.query
        .join(Warehouse, Stock.warehouse_id == Warehouse.id)
        .join(Product, Stock.product_id == Product.id)
        .options(db.joinedload(Stock.product), db.joinedload(Stock.warehouse))
        .order_by(Warehouse.name.asc(), Product.name.asc())
    )
    if owner_id is not None:
        q = q.filter(Warehouse.owner_id == owner_id)

    rows = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = ["Warehouse", "Product", "SKU", "Category", "Qty"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for s in rows:
        ws.append([
            (s.warehouse.name if s.warehouse else ""),
            (s.product.name if s.product else ""),
            (s.product.sku if s.product else ""),
            (s.product.category if s.product else ""),
            int(s.quantity or 0),
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="stock.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/transactions.pdf")
@login_required
@roles_required("Admin / Owner", "Warehouse Manager", "Developer", "Sales Agent")
def transactions_pdf():
    owner_id = _get_owner_id()
    if not _require_owner_context(owner_id):
        return redirect(url_for("reports.reports_home"))

    q = _apply_tx_filters(_transactions_query(owner_id))
    txns = q.limit(200).all()  # keep PDF readable

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    width, height = A4

    y = height - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Transactions Report")
    y -= 25

    c.setFont("Helvetica", 9)
    c.drawString(40, y, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    y -= 20

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Date")
    c.drawString(130, y, "Type")
    c.drawString(180, y, "Partner")
    c.drawString(300, y, "Warehouse")
    c.drawString(420, y, "Total")
    y -= 12
    c.line(40, y, width - 40, y)
    y -= 14

    c.setFont("Helvetica", 9)
    for t in txns:
        total = sum((i.total_price or 0) for i in (t.items or []))
        c.drawString(40, y, t.date.strftime("%Y-%m-%d") if t.date else "")
        c.drawString(130, y, t.type or "")
        c.drawString(180, y, (t.partner.name if t.partner else "")[:18])
        c.drawString(300, y, (t.warehouse.name if t.warehouse else "")[:18])
        c.drawRightString(width - 40, y, f"{float(total):.2f}")
        y -= 14

        if y < 60:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica", 9)

    c.save()
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="transactions.pdf",
        mimetype="application/pdf",
    )
